import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from evaluator import generate_review


class GenerateReviewWorkbookTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory(dir=generate_review.REPO_ROOT)
        self.addCleanup(self.tempdir.cleanup)
        self.results_root = Path(self.tempdir.name) / "results"
        self.results_root.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def _ground_truth_rows() -> dict[int, dict[str, str]]:
        return {
            10: {
                "bug_id": "10",
                "app_name": "Test App 10",
                "HC_LP Desc": "HC_LP description",
                "MC_MP Desc": "MC_MP description",
                "S2R_ground_truth": "1. First step\n2. Second step",
                "info_elements_gt": "gt info 10",
            },
            20: {
                "bug_id": "20",
                "app_name": "Test App 20",
                "HC_HP Desc": "HC_HP description",
                "S2R_ground_truth": "1. Only step",
                "info_elements_gt": "gt info 20",
            },
        }

    def _write_evaluation(self, directory: Path, filename: str, payload: dict) -> None:
        directory.mkdir(parents=True, exist_ok=True)
        (directory / filename).write_text(json.dumps(payload, indent=2), encoding="utf-8")

    @staticmethod
    def _evaluation_payload(
        *,
        bug_id: int,
        agent_version: str,
        description_level: str,
        total_tokens_consumed=None,
        total_wall_clock_seconds=None,
        total_turn_processing_seconds=None,
        total_conversation_turns=None,
        s2r_judge=None,
        info_elements_judge=None,
    ) -> dict:
        return {
            "log_path": f"logs/{agent_version}/bug{bug_id}_{description_level}.log",
            "bug_id": bug_id,
            "description_level": description_level,
            "agent_version": agent_version,
            "app_name": f"Test App {bug_id}",
            "final_report": {
                "title": f"Bug {bug_id}",
                "observed_behavior": "Observed",
                "expected_behavior": "Expected",
            },
            "ground_truth": {
                "bug_id": str(bug_id),
                "app_name": f"Test App {bug_id}",
                "info_elements_gt": f"gt info {bug_id}",
                "S2R_ground_truth": (
                    "1. First step\n2. Second step" if bug_id == 10 else "1. Only step"
                ),
            },
            "recomputed_info_elements": {
                "buggy_behavior": "Observed",
                "triggering_gui_interactions": "Tap button",
                "triggering_screen_reference": "Main screen",
                "correct_behavior": "Expected",
            },
            "info_elements_judge": info_elements_judge
            or {
                "buggy_behavior": "Correct",
                "triggering_gui_interactions": "Missing",
                "triggering_screen_reference": "Ambiguous",
                "correct_behavior": "Incorrect",
            },
            "s2r_judge": s2r_judge
            or [
                {
                    "generated_step": "Open app",
                    "label": "CS",
                    "matched_gt_step": "1. First step",
                },
                {
                    "generated_step": "Tap button",
                    "label": "ES",
                    "matched_gt_step": "",
                },
            ],
            "total_tokens_consumed": total_tokens_consumed,
            "total_wall_clock_seconds": total_wall_clock_seconds,
            "total_turn_processing_seconds": total_turn_processing_seconds,
            "total_conversation_turns": total_conversation_turns,
        }

    def test_rebuild_manual_review_workbook_adds_summary_sheet_and_hidden_gt_helper(self):
        version_dir = self.results_root / "VTest"
        self._write_evaluation(
            version_dir,
            "bug10_HC_LP.evaluation.json",
            self._evaluation_payload(
                bug_id=10,
                agent_version="VTest",
                description_level="HC_LP",
                total_tokens_consumed=100,
                total_wall_clock_seconds=12.5,
                total_turn_processing_seconds=8.0,
                total_conversation_turns=3,
            ),
        )

        with patch.object(generate_review, "RESULTS_ROOT", self.results_root), patch.object(
            generate_review, "load_ground_truth_rows", return_value=self._ground_truth_rows()
        ):
            workbook_path = generate_review.rebuild_manual_review_workbook("VTest")

        workbook = load_workbook(workbook_path)
        self.assertEqual(
            workbook.sheetnames,
            ["S2R Review", "Info Elements Review", "Summary"],
        )

        s2r_sheet = workbook["S2R Review"]
        self.assertEqual(s2r_sheet["N1"].value, "F1")
        self.assertEqual(s2r_sheet["O1"].value, "GT Step Count")
        self.assertTrue(s2r_sheet.column_dimensions["O"].hidden)
        self.assertEqual(s2r_sheet["O2"].value, 2)
        self.assertIsNone(s2r_sheet["O3"].value)
        self.assertEqual(s2r_sheet["L2"].value, '=IFERROR(COUNTIF(K2:K3,"CS")/COUNTA(K2:K3),"")')
        self.assertEqual(s2r_sheet["M2"].value, '=COUNTIF(K2:K3,"CS")/2')
        self.assertEqual(s2r_sheet["N2"].value, '=IFERROR(2*(L2*M2)/(L2+M2),"")')

        summary_sheet = workbook["Summary"]
        headers = [summary_sheet.cell(row=1, column=column).value for column in range(1, 17)]
        self.assertEqual(
            headers,
            [
                "agent_version",
                "s2r_cs_count",
                "s2r_ms_count",
                "s2r_es_count",
                "average_s2r_precision",
                "average_s2r_recall",
                "average_s2r_f1",
                "info_correct_count",
                "info_incomplete_count",
                "info_ambiguous_count",
                "info_missing_count",
                "info_incorrect_count",
                "average_total_tokens_consumed_per_conv",
                "average_wall_clock_seconds_per_conv",
                "average_turn_processing_seconds_per_conv",
                "average_conversation_turns",
            ],
        )
        self.assertEqual(summary_sheet["A2"].value, "VTest")
        self.assertEqual(
            summary_sheet["B2"].value,
            '=COUNTIF(\'S2R Review\'!$K:$K,"CS")',
        )
        self.assertEqual(
            summary_sheet["C2"].value,
            '=SUM(\'S2R Review\'!$O:$O)-B2',
        )
        self.assertEqual(
            summary_sheet["D2"].value,
            '=COUNTIF(\'S2R Review\'!$K:$K,"ES")',
        )
        self.assertEqual(
            summary_sheet["E2"].value,
            '=IFERROR(SUM(\'S2R Review\'!$L$2:$L$1048576)/COUNTIF(\'S2R Review\'!$L$2:$L$1048576,"<>"),"")',
        )
        self.assertEqual(
            summary_sheet["F2"].value,
            '=IFERROR(SUM(\'S2R Review\'!$M$2:$M$1048576)/COUNTIF(\'S2R Review\'!$M$2:$M$1048576,"<>"),"")',
        )
        self.assertEqual(
            summary_sheet["G2"].value,
            '=IFERROR(SUM(\'S2R Review\'!$N$2:$N$1048576)/COUNTIF(\'S2R Review\'!$N$2:$N$1048576,"<>"),"")',
        )
        self.assertEqual(
            summary_sheet["H2"].value,
            '=COUNTIF(\'Info Elements Review\'!$I:$I,"Correct")+COUNTIF(\'Info Elements Review\'!$J:$J,"Correct")+COUNTIF(\'Info Elements Review\'!$K:$K,"Correct")+COUNTIF(\'Info Elements Review\'!$L:$L,"Correct")',
        )
        self.assertEqual(summary_sheet["M2"].value, 100.0)
        self.assertEqual(summary_sheet["N2"].value, 12.5)
        self.assertEqual(summary_sheet["O2"].value, 8.0)
        self.assertEqual(summary_sheet["P2"].value, 3.0)

    def test_rebuild_manual_review_workbook_uses_single_summary_row_and_ignores_null_json_metrics(self):
        version_dir = self.results_root / "VTest"
        self._write_evaluation(
            version_dir,
            "bug10_HC_LP.evaluation.json",
            self._evaluation_payload(
                bug_id=10,
                agent_version="AV1",
                description_level="HC_LP",
                total_tokens_consumed=100,
                total_wall_clock_seconds=12.0,
                total_turn_processing_seconds=7.0,
                total_conversation_turns=2,
            ),
        )
        self._write_evaluation(
            version_dir,
            "bug20_HC_HP.evaluation.json",
            self._evaluation_payload(
                bug_id=20,
                agent_version="AV1",
                description_level="HC_HP",
                total_tokens_consumed=None,
                total_wall_clock_seconds=18.0,
                total_turn_processing_seconds=9.0,
                total_conversation_turns=None,
                s2r_judge=[
                    {
                        "generated_step": "Do step",
                        "label": "CS",
                        "matched_gt_step": "1. Only step",
                    }
                ],
                info_elements_judge={
                    "buggy_behavior": "Incomplete",
                    "triggering_gui_interactions": "Correct",
                    "triggering_screen_reference": "Correct",
                    "correct_behavior": "Missing",
                },
            ),
        )
        self._write_evaluation(
            version_dir,
            "bug10_MC_MP.evaluation.json",
            self._evaluation_payload(
                bug_id=10,
                agent_version="BV2",
                description_level="MC_MP",
                total_tokens_consumed=None,
                total_wall_clock_seconds=None,
                total_turn_processing_seconds=None,
                total_conversation_turns=None,
            ),
        )

        with patch.object(generate_review, "RESULTS_ROOT", self.results_root), patch.object(
            generate_review, "load_ground_truth_rows", return_value=self._ground_truth_rows()
        ):
            workbook_path = generate_review.rebuild_manual_review_workbook("VTest")

        workbook = load_workbook(workbook_path)
        summary_sheet = workbook["Summary"]

        self.assertEqual(summary_sheet.max_row, 2)
        self.assertEqual(summary_sheet["A2"].value, "VTest")
        self.assertEqual(summary_sheet["B2"].value, '=COUNTIF(\'S2R Review\'!$K:$K,"CS")')
        self.assertEqual(summary_sheet["C2"].value, '=SUM(\'S2R Review\'!$O:$O)-B2')
        self.assertEqual(summary_sheet["D2"].value, '=COUNTIF(\'S2R Review\'!$K:$K,"ES")')
        self.assertEqual(
            summary_sheet["G2"].value,
            '=IFERROR(SUM(\'S2R Review\'!$N$2:$N$1048576)/COUNTIF(\'S2R Review\'!$N$2:$N$1048576,"<>"),"")',
        )
        self.assertEqual(summary_sheet["M2"].value, 100.0)
        self.assertEqual(summary_sheet["N2"].value, 15.0)
        self.assertEqual(summary_sheet["O2"].value, 8.0)
        self.assertEqual(summary_sheet["P2"].value, 2.0)

        s2r_sheet = workbook["S2R Review"]
        self.assertEqual(s2r_sheet["O2"].value, 2)
        self.assertEqual(s2r_sheet["O5"].value, 2)
        self.assertEqual(s2r_sheet["O8"].value, 1)
        self.assertIsNone(s2r_sheet["A4"].value)
        self.assertIsNone(s2r_sheet["O4"].value)


if __name__ == "__main__":
    unittest.main()
