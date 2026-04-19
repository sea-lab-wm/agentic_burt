from __future__ import annotations

"""Workbook-generation helpers for manual review of evaluator outputs."""

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill

from evaluator.parsing import REPO_ROOT, load_ground_truth_rows


RESULTS_ROOT = REPO_ROOT / "Results"
MANUAL_REVIEW_WORKBOOK = "manual_review.xlsx"
SEPARATOR_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")


def rebuild_manual_review_workbook(agent_version: str) -> Path:
    """Rebuild the combined manual-review workbook for one agent version."""
    version_dir = RESULTS_ROOT / agent_version
    workbook_path = version_dir / MANUAL_REVIEW_WORKBOOK
    evaluation_paths = sorted(version_dir.glob("*.evaluation.json"))
    evaluation_results = [
        json.loads(evaluation_path.read_text(encoding="utf-8"))
        for evaluation_path in evaluation_paths
    ]
    ground_truth_rows = load_ground_truth_rows()

    workbook = Workbook()
    s2r_sheet = workbook.active
    s2r_sheet.title = "S2R Review"
    _populate_s2r_review_sheet(s2r_sheet, evaluation_results, ground_truth_rows)

    info_element_sheet = workbook.create_sheet("Info Elements Review")
    _populate_info_elements_review_sheet(info_element_sheet, evaluation_results, ground_truth_rows)

    summary_sheet = workbook.create_sheet("Summary")
    _populate_summary_sheet(summary_sheet, agent_version, evaluation_results)

    workbook.save(workbook_path)
    return workbook_path


def _populate_s2r_review_sheet(
    sheet: Any, evaluation_results: list[dict[str, Any]], ground_truth_rows: dict[int, dict[str, str]]
) -> None:
    """Populate the S2R review sheet for the combined manual-review workbook."""
    headers = [
        "bug_id",
        "app_name",
        "agent_version",
        "description_level",
        "description_text",
        "full_bug_report",
        "s2r_gt",
        "agent_generated_steps",
        "LLM_evaluation",
        "Matched_GT_by_LLM",
        "human_evaluation",
        "Precision",
        "Recall",
        "F1",
        "GT Step Count",
    ]
    sheet.append(headers)
    sheet.freeze_panes = "A2"

    for column, width in {
        "A": 10,
        "B": 22,
        "C": 16,
        "D": 18,
        "E": 60,
        "F": 60,
        "G": 60,
        "H": 60,
        "I": 16,
        "J": 60,
        "K": 18,
        "L": 14,
        "M": 14,
        "N": 14,
        "O": 14,
    }.items():
        sheet.column_dimensions[column].width = width
    sheet.column_dimensions["O"].hidden = True

    next_row = 2
    for evaluation_index, result in enumerate(evaluation_results):
        bug_id = result.get("bug_id")
        description_level = result.get("description_level")
        gt_row = ground_truth_rows.get(bug_id) if isinstance(bug_id, int) else None
        s2r_rows = result.get("s2r_judge")
        block_rows = s2r_rows if isinstance(s2r_rows, list) and s2r_rows else [None]
        block_start = next_row
        block_end = block_start + len(block_rows) - 1
        gt_text = ((result.get("ground_truth") or {}).get("S2R_ground_truth") or "").strip()
        gt_count = _count_numbered_steps(gt_text)

        for index, step in enumerate(block_rows):
            row_number = block_start + index
            row_values = [
                bug_id if index == 0 else "",
                result.get("app_name") if index == 0 else "",
                result.get("agent_version") if index == 0 else "",
                description_level if index == 0 else "",
                _lookup_description_text(gt_row, description_level) if index == 0 else "",
                _build_full_bug_report_text(result) if index == 0 else "",
                gt_text if index == 0 else "",
                (step or {}).get("generated_step", ""),
                (step or {}).get("label", ""),
                (step or {}).get("matched_gt_step", ""),
                (step or {}).get("label", ""),
                _build_precision_formula(block_start, block_end) if index == 0 else "",
                _build_recall_formula(block_start, block_end, gt_count) if index == 0 else "",
                _build_f1_formula(row_number) if index == 0 else "",
                gt_count if index == 0 else "",
            ]
            sheet.append(row_values)
            for cell in sheet[row_number]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        next_row = block_end + 1
        if evaluation_index < len(evaluation_results) - 1:
            _append_separator_row(sheet, len(headers), next_row)
            next_row += 1

def _populate_info_elements_review_sheet(
    sheet: Any, evaluation_results: list[dict[str, Any]], ground_truth_rows: dict[int, dict[str, str]]
) -> None:
    """Populate the information-elements review sheet for the combined manual-review workbook."""
    headers = [
        "bug_id",
        "app_name",
        "agent_version",
        "description_level",
        "description_text",
        "full_generated_bug_report",
        "info_elements_gt",
        "agent_generated_info_elements",
        "buggy_behavior_grade",
        "triggering_gui_interactions_grade",
        "triggering_screen_reference_grade",
        "correct_behavior_grade",
    ]
    sheet.append(headers)
    sheet.freeze_panes = "A2"

    for column, width in {
        "A": 10,
        "B": 22,
        "C": 16,
        "D": 18,
        "E": 60,
        "F": 60,
        "G": 60,
        "H": 60,
        "I": 20,
        "J": 26,
        "K": 26,
        "L": 20,
    }.items():
        sheet.column_dimensions[column].width = width

    next_row = 2
    for evaluation_index, result in enumerate(evaluation_results):
        bug_id = result.get("bug_id")
        description_level = result.get("description_level")
        gt_row = ground_truth_rows.get(bug_id) if isinstance(bug_id, int) else None
        info_judge = result.get("info_elements_judge")
        row_values = [
            bug_id,
            result.get("app_name"),
            result.get("agent_version"),
            description_level,
            _lookup_description_text(gt_row, description_level),
            _build_full_bug_report_text(result),
            ((result.get("ground_truth") or {}).get("info_elements_gt") or "").strip(),
            _format_agent_generated_info_elements(result.get("recomputed_info_elements")),
            info_judge.get("buggy_behavior", "") if isinstance(info_judge, dict) else "",
            info_judge.get("triggering_gui_interactions", "") if isinstance(info_judge, dict) else "",
            info_judge.get("triggering_screen_reference", "") if isinstance(info_judge, dict) else "",
            info_judge.get("correct_behavior", "") if isinstance(info_judge, dict) else "",
        ]
        sheet.append(row_values)
        for cell in sheet[next_row]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        next_row += 1
        if evaluation_index < len(evaluation_results) - 1:
            _append_separator_row(sheet, len(headers), next_row)
            next_row += 1


def _populate_summary_sheet(
    sheet: Any, agent_version: str, evaluation_results: list[dict[str, Any]]
) -> None:
    """Populate the workbook summary sheet for the current agent version."""
    headers = [
        "agent_version",
        "s2r_cs_count",
        "s2r_ms_count",
        "s2r_es_count",
        "average_s2r_precision",
        "average_s2r_recall",
        "average_s2r_f1",
        "info_correct_count",
        "info_incomplete_count",
        "info_ambiguous_count",
        "info_missing_count",
        "info_incorrect_count",
        "average_total_tokens_consumed_per_conv",
        "average_wall_clock_seconds_per_conv",
        "average_turn_processing_seconds_per_conv",
        "average_conversation_turns",
    ]
    sheet.append(headers)
    sheet.freeze_panes = "A2"

    for column, width in {
        "A": 18,
        "B": 14,
        "C": 14,
        "D": 14,
        "E": 22,
        "F": 20,
        "G": 18,
        "H": 18,
        "I": 20,
        "J": 19,
        "K": 18,
        "L": 20,
        "M": 34,
        "N": 32,
        "O": 34,
        "P": 26,
    }.items():
        sheet.column_dimensions[column].width = width

    token_values: list[float] = []
    wall_clock_values: list[float] = []
    turn_processing_values: list[float] = []
    turns_values: list[float] = []
    for result in evaluation_results:
        _append_numeric_metric(token_values, result.get("total_tokens_consumed"))
        _append_numeric_metric(wall_clock_values, result.get("total_wall_clock_seconds"))
        _append_numeric_metric(
            turn_processing_values, result.get("total_turn_processing_seconds")
        )
        _append_numeric_metric(turns_values, result.get("total_conversation_turns"))

    row_number = 2
    row_values = [
        agent_version,
        _build_s2r_label_count_formula("CS"),
        _build_s2r_ms_count_formula(row_number),
        _build_s2r_label_count_formula("ES"),
        _build_sheet_average_formula("S2R Review", "L"),
        _build_sheet_average_formula("S2R Review", "M"),
        _build_sheet_average_formula("S2R Review", "N"),
        _build_info_label_count_formula("Correct"),
        _build_info_label_count_formula("Incomplete"),
        _build_info_label_count_formula("Ambiguous"),
        _build_info_label_count_formula("Missing"),
        _build_info_label_count_formula("Incorrect"),
        _average_or_none(token_values),
        _average_or_none(wall_clock_values),
        _average_or_none(turn_processing_values),
        _average_or_none(turns_values),
    ]
    sheet.append(row_values)


def _append_numeric_metric(values: list[float], candidate: Any) -> None:
    """Append one numeric metric to an accumulator list when it is valid."""
    if isinstance(candidate, (int, float)) and not isinstance(candidate, bool):
        values.append(float(candidate))


def _average_or_none(values: list[float]) -> float | None:
    """Return the arithmetic mean or None for empty inputs."""
    if not values:
        return None
    return sum(values) / len(values)


def _build_s2r_label_count_formula(label: str) -> str:
    """Count one S2R human-evaluation label across the workbook."""
    return f'=COUNTIF(\'S2R Review\'!$K:$K,"{label}")'


def _build_s2r_ms_count_formula(row_number: int) -> str:
    """Compute workbook-level missing S2R count as GT total minus CS count."""
    return f'=SUM(\'S2R Review\'!$O:$O)-B{row_number}'


def _build_sheet_average_formula(sheet_name: str, value_column: str) -> str:
    """Build an Excel formula that averages non-blank values in one column."""
    return (
        f'=IFERROR(SUM(\'{sheet_name}\'!${value_column}$2:${value_column}$1048576)/'
        f'COUNTIF(\'{sheet_name}\'!${value_column}$2:${value_column}$1048576,"<>"),"")'
    )


def _build_info_label_count_formula(label: str) -> str:
    """Count one info-element label across all four review grade columns."""
    return (
        f'=COUNTIF(\'Info Elements Review\'!$I:$I,"{label}")+'
        f'COUNTIF(\'Info Elements Review\'!$J:$J,"{label}")+'
        f'COUNTIF(\'Info Elements Review\'!$K:$K,"{label}")+'
        f'COUNTIF(\'Info Elements Review\'!$L:$L,"{label}")'
    )

def _lookup_description_text(gt_row: dict[str, str] | None, description_level: str | None) -> str:
    """Resolve the input description text for the bug/description pair from the dev CSV."""
    if not gt_row or not description_level:
        return ""
    column_name = f"{description_level} Desc"
    return (gt_row.get(column_name) or "").strip()


def _build_full_bug_report_text(result: dict[str, Any]) -> str:
    """Compose a display-friendly bug-report text block for workbook cells."""
    final_report = result.get("final_report") or {}
    parts = [
        ("Title", final_report.get("title")),
        ("Observed Behavior", final_report.get("observed_behavior")),
        ("Expected Behavior", final_report.get("expected_behavior")),
    ]
    return "\n\n".join(f"{label}: {value}" for label, value in parts if value)


def _format_agent_generated_info_elements(info_elements: dict[str, str] | None) -> str:
    """Render generated information elements into a manual-review text block."""
    if not info_elements:
        return ""

    parts = [
        ("Buggy Behavior", info_elements.get("buggy_behavior")),
        ("Triggering GUI Interactions", info_elements.get("triggering_gui_interactions")),
        ("Triggering Screen Reference", info_elements.get("triggering_screen_reference")),
        ("Correct Behavior", info_elements.get("correct_behavior")),
    ]
    return "\n".join(f"{label}: {value}" for label, value in parts if value)


def _build_precision_formula(start_row: int, end_row: int) -> str:
    """Build the per-block precision formula from human-evaluation labels."""
    return f'=IFERROR(COUNTIF(K{start_row}:K{end_row},"CS")/COUNTA(K{start_row}:K{end_row}),"")'


def _build_recall_formula(start_row: int, end_row: int, gt_count: int) -> str:
    """Build the per-block recall formula from human-evaluation labels."""
    if gt_count <= 0:
        return ""
    return f'=COUNTIF(K{start_row}:K{end_row},"CS")/{gt_count}'


def _build_f1_formula(row_number: int) -> str:
    """Build the per-block F1 formula from the block precision and recall cells."""
    return f'=IFERROR(2*(L{row_number}*M{row_number})/(L{row_number}+M{row_number}),"")'


def _count_numbered_steps(steps_text: str) -> int:
    """Count numbered S2R lines in the ground-truth CSV format."""
    return sum(
        1
        for line in steps_text.splitlines()
        if line.strip() and line.lstrip().split(".", 1)[0].isdigit()
    )


def _append_separator_row(sheet: Any, column_count: int, row_number: int) -> None:
    """Insert a blank gray spacer row between review blocks."""
    sheet.append([""] * column_count)
    for cell in sheet[row_number]:
        cell.fill = SEPARATOR_FILL
