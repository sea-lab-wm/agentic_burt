from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from langchain_openai import ChatOpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill

import config
from evaluator.judges import (
    InfoElementsJudgeResult,
    S2RJudgeResult,
    extract_information_elements_from_OB_EB,
    judge_information_elements,
    judge_s2r,
)
from evaluator.parsing import REPO_ROOT, build_log_context, discover_log_paths, load_ground_truth_rows


RESULTS_ROOT = REPO_ROOT / "Results"
S2R_REVIEW_WORKBOOK = "s2r_manual_review.xlsx"
INFO_ELEMENTS_REVIEW_WORKBOOK = "information_elements_manual_review.xlsx"
SEPARATOR_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")


def parse_args() -> argparse.Namespace:
    """Parse evaluator CLI arguments."""
    parser = argparse.ArgumentParser(description="Evaluate BURT logs with LLM-as-judge prompts.")
    parser.add_argument("paths", nargs="+", help="One or more log files or directories.")
    parser.add_argument(
        "--model",
        default=config.MODEL_NAME,
        help=f"Judge model name. Defaults to {config.MODEL_NAME}.",
    )
    return parser.parse_args()


def main() -> None:
    """Run the full evaluator pipeline for the requested log inputs."""
    load_dotenv()
    args = parse_args()

    log_paths = discover_log_paths(args.paths)
    if not log_paths:
        raise SystemExit("No log files were found for the provided inputs.")

    ground_truth_rows = load_ground_truth_rows()
    model = ChatOpenAI(model=args.model)

    results_by_version: dict[str, list[Path]] = {}
    for log_path in log_paths:
        # Each log is evaluated independently and persisted immediately so a
        # partial run still leaves behind useful artifacts.
        result = evaluate_log(log_path=log_path, model=model, ground_truth_rows=ground_truth_rows)
        output_path = write_evaluation_result(result)
        results_by_version.setdefault(result["agent_version"], []).append(output_path)

    for agent_version in results_by_version:
        rebuild_summary_csv(agent_version)
        rebuild_s2r_review_workbook(agent_version)
        rebuild_info_elements_review_workbook(agent_version)


def evaluate_log(log_path: Path, model: Any, ground_truth_rows: dict[int, dict[str, str]]) -> dict[str, Any]:
    """Evaluate one parsed BURT log and return the persisted JSON payload."""
    context = build_log_context(log_path, ground_truth_rows)
    timestamp = datetime.now(timezone.utc).isoformat()

    result: dict[str, Any] = {
        "log_path": context["log_path"],
        "bug_id": context["bug_id"],
        "description_level": context["description_level"],
        "agent_version": context["agent_version"],
        "app_name": context.get("app_name"),
        "title": (context.get("full_report") or {}).get("title"),
        "full_report": context.get("full_report"),
        "ground_truth": context.get("ground_truth"),
        "recomputed_info_elements": None,
        "info_elements_judge": None,
        "s2r_judge": None,
        "judge_model": getattr(model, "model_name", None) or getattr(model, "model", None),
        "evaluated_at": timestamp,
        "status": "ok",
        "error": None,
        "parse_status": context["parse_status"],
        "parse_error": context["parse_error"],
    }

    if context["parse_status"] != "ok":
        # Parsing failures are written as result artifacts instead of raising so
        # multi-log runs can continue and the failure remains inspectable.
        result["status"] = "parse_error"
        return result

    full_report = context["full_report"] or {}
    ground_truth = context.get("ground_truth") or {}

    info_judge_result: InfoElementsJudgeResult | None = None
    s2r_judge_result: S2RJudgeResult | None = None

    try:
        # Recompute the information elements from the final report so judging is
        # based on a fresh extraction pass, not only the values captured in the log.
        result["recomputed_info_elements"] = extract_information_elements_from_OB_EB(
            observed_behavior=full_report.get("observed_behavior", ""),
            expected_behavior=full_report.get("expected_behavior", ""),
            model=model,
        )
    except Exception as exc:
        result["status"] = "judge_error"
        result["error"] = f"Failed to recompute info elements: {exc}"

    if result["recomputed_info_elements"] and ground_truth.get("info_elements_gt"):
        try:
            info_judge_result = judge_information_elements(
                generated_info_elements=result["recomputed_info_elements"],
                ground_truth_info_elements=ground_truth["info_elements_gt"],
                model=model,
            )
            result["info_elements_judge"] = info_judge_result.model_dump()
        except Exception as exc:
            result["status"] = "judge_error"
            result["error"] = _append_error(result["error"], f"Information-elements judge failed: {exc}")
    else:
        result["info_elements_judge"] = {
            "status": "skipped",
            "reason": "Missing recomputed info elements or info_elements_gt.",
        }

    if full_report.get("steps_to_reproduce") and ground_truth.get("S2R_ground_truth"):
        try:
            s2r_judge_result = judge_s2r(
                generated_s2rs=full_report["steps_to_reproduce"],
                ground_truth_s2r=ground_truth["S2R_ground_truth"],
                model=model,
            )
            result["s2r_judge"] = [step.model_dump() for step in s2r_judge_result.steps]
        except Exception as exc:
            result["status"] = "judge_error"
            result["error"] = _append_error(result["error"], f"S2R judge failed: {exc}")
    else:
        result["s2r_judge"] = {
            "status": "skipped",
            "reason": "Missing generated S2R or S2R_ground_truth.",
        }

    return result


def write_evaluation_result(result: dict[str, Any]) -> Path:
    """Write one evaluation JSON under Results/<agent_version>/."""
    version_dir = RESULTS_ROOT / result["agent_version"]
    version_dir.mkdir(parents=True, exist_ok=True)
    output_path = version_dir / f"{Path(result['log_path']).stem}.evaluation.json"
    output_path.write_text(json.dumps(result, indent=2), encoding="utf-8")
    return output_path


def rebuild_summary_csv(agent_version: str) -> Path:
    """Regenerate the version-local summary CSV from evaluation JSON files."""
    version_dir = RESULTS_ROOT / agent_version
    summary_path = version_dir / "summary.csv"
    evaluation_paths = sorted(version_dir.glob("*.evaluation.json"))

    fieldnames = [
        "bug_id",
        "description_level",
        "agent_version",
        "log_file",
        "log_path",
        "title",
        "status",
        "parse_status",
        "error",
    ]

    with summary_path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        writer.writeheader()

        for evaluation_path in evaluation_paths:
            result = json.loads(evaluation_path.read_text(encoding="utf-8"))
            writer.writerow(
                {
                    "bug_id": result.get("bug_id"),
                    "description_level": result.get("description_level"),
                    "agent_version": result.get("agent_version"),
                    "log_file": Path(result.get("log_path", "")).name,
                    "log_path": result.get("log_path"),
                    "title": result.get("title"),
                    "status": result.get("status"),
                    "parse_status": result.get("parse_status"),
                    "error": result.get("error"),
                }
            )

    return summary_path


def rebuild_s2r_review_workbook(agent_version: str) -> Path:
    """Build a manual-review workbook for S2R judge outputs."""
    version_dir = RESULTS_ROOT / agent_version
    workbook_path = version_dir / S2R_REVIEW_WORKBOOK
    evaluation_paths = sorted(version_dir.glob("*.evaluation.json"))
    ground_truth_rows = load_ground_truth_rows()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "S2R Review"
    headers = [
        "bug_id",
        "app_name",
        "agent_version",
        "description_level",
        "description_text",
        "full_bug_report",
        "s2r_gt",
        "agent_generated_steps",
        "LLM_evaluation",
        "Matched_GT_by_LLM",
        "human_evaluation",
        "Precision",
        "Recall",
    ]
    sheet.append(headers)
    sheet.freeze_panes = "A2"

    for column, width in {
        "A": 10,
        "B": 22,
        "C": 16,
        "D": 18,
        "E": 60,
        "F": 60,
        "G": 60,
        "H": 60,
        "I": 16,
        "J": 60,
        "K": 18,
        "L": 14,
        "M": 14,
    }.items():
        sheet.column_dimensions[column].width = width

    next_row = 2
    for evaluation_index, evaluation_path in enumerate(evaluation_paths):
        result = json.loads(evaluation_path.read_text(encoding="utf-8"))
        bug_id = result.get("bug_id")
        description_level = result.get("description_level")
        gt_row = ground_truth_rows.get(bug_id) if isinstance(bug_id, int) else None
        s2r_rows = result.get("s2r_judge")
        block_rows = s2r_rows if isinstance(s2r_rows, list) and s2r_rows else [None]
        block_start = next_row
        block_end = block_start + len(block_rows) - 1
        gt_text = ((result.get("ground_truth") or {}).get("S2R_ground_truth") or "").strip()
        gt_count = _count_numbered_steps(gt_text)

        for index, step in enumerate(block_rows):
            row_number = block_start + index
            row_values = [
                bug_id if index == 0 else "",
                result.get("app_name") if index == 0 else "",
                result.get("agent_version") if index == 0 else "",
                description_level if index == 0 else "",
                _lookup_description_text(gt_row, description_level) if index == 0 else "",
                _build_full_bug_report_text(result) if index == 0 else "",
                gt_text if index == 0 else "",
                (step or {}).get("generated_step", ""),
                (step or {}).get("label", ""),
                (step or {}).get("matched_gt_step", ""),
                (step or {}).get("label", ""),
                _build_precision_formula(block_start, block_end) if index == 0 else "",
                _build_recall_formula(block_start, block_end, gt_count) if index == 0 else "",
            ]
            sheet.append(row_values)
            for cell in sheet[row_number]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        next_row = block_end + 1
        if evaluation_index < len(evaluation_paths) - 1:
            _append_separator_row(sheet, len(headers), next_row)
            next_row += 1

    workbook.save(workbook_path)
    return workbook_path


def rebuild_info_elements_review_workbook(agent_version: str) -> Path:
    """Build a manual-review workbook for information-elements judge outputs."""
    version_dir = RESULTS_ROOT / agent_version
    workbook_path = version_dir / INFO_ELEMENTS_REVIEW_WORKBOOK
    evaluation_paths = sorted(version_dir.glob("*.evaluation.json"))
    ground_truth_rows = load_ground_truth_rows()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Info Elements Review"
    headers = [
        "bug_id",
        "app_name",
        "agent_version",
        "description_level",
        "description_text",
        "full_generated_bug_report",
        "info_elements_gt",
        "agent_generated_info_elements",
        "buggy_behavior_grade",
        "triggering_gui_interactions_grade",
        "triggering_screen_reference_grade",
        "correct_behavior_grade",
    ]
    sheet.append(headers)
    sheet.freeze_panes = "A2"

    for column, width in {
        "A": 10,
        "B": 22,
        "C": 16,
        "D": 18,
        "E": 60,
        "F": 60,
        "G": 60,
        "H": 60,
        "I": 20,
        "J": 26,
        "K": 26,
        "L": 20,
    }.items():
        sheet.column_dimensions[column].width = width

    next_row = 2
    for evaluation_index, evaluation_path in enumerate(evaluation_paths):
        result = json.loads(evaluation_path.read_text(encoding="utf-8"))
        bug_id = result.get("bug_id")
        description_level = result.get("description_level")
        gt_row = ground_truth_rows.get(bug_id) if isinstance(bug_id, int) else None
        info_judge = result.get("info_elements_judge")
        row_values = [
            bug_id,
            result.get("app_name"),
            result.get("agent_version"),
            description_level,
            _lookup_description_text(gt_row, description_level),
            _build_full_bug_report_text(result),
            ((result.get("ground_truth") or {}).get("info_elements_gt") or "").strip(),
            _format_agent_generated_info_elements(result.get("recomputed_info_elements")),
            info_judge.get("buggy_behavior", "") if isinstance(info_judge, dict) else "",
            info_judge.get("triggering_gui_interactions", "") if isinstance(info_judge, dict) else "",
            info_judge.get("triggering_screen_reference", "") if isinstance(info_judge, dict) else "",
            info_judge.get("correct_behavior", "") if isinstance(info_judge, dict) else "",
        ]
        sheet.append(row_values)
        for cell in sheet[next_row]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        next_row += 1
        if evaluation_index < len(evaluation_paths) - 1:
            _append_separator_row(sheet, len(headers), next_row)
            next_row += 1

    workbook.save(workbook_path)
    return workbook_path


def _lookup_description_text(gt_row: dict[str, str] | None, description_level: str | None) -> str:
    """Resolve the input description text for the bug/description pair from the dev CSV."""
    if not gt_row or not description_level:
        return ""
    column_name = f"{description_level} Desc"
    return (gt_row.get(column_name) or "").strip()


def _build_full_bug_report_text(result: dict[str, Any]) -> str:
    """Compose a consistent generated bug report text block."""
    full_report = result.get("full_report") or {}
    parts = [
        ("Title", full_report.get("title")),
        ("Observed Behavior", full_report.get("observed_behavior")),
        ("Expected Behavior", full_report.get("expected_behavior")),
    ]
    return "\n\n".join(f"{label}: {value}" for label, value in parts if value)


def _format_agent_generated_info_elements(info_elements: dict[str, str] | None) -> str:
    """Render generated information elements into a manual-review text block."""
    if not info_elements:
        return ""

    parts = [
        ("Buggy Behavior", info_elements.get("buggy_behavior")),
        ("Triggering GUI Interactions", info_elements.get("triggering_gui_interactions")),
        ("Triggering Screen Reference", info_elements.get("triggering_screen_reference")),
        ("Correct Behavior", info_elements.get("correct_behavior")),
    ]
    return "\n".join(f"{label}: {value}" for label, value in parts if value)


def _build_precision_formula(start_row: int, end_row: int) -> str:
    """Build the per-block precision formula from the human evaluation column."""
    return f'=IFERROR(COUNTIF(K{start_row}:K{end_row},"CS")/COUNTA(K{start_row}:K{end_row}),"")'


def _build_recall_formula(start_row: int, end_row: int, gt_count: int) -> str:
    """Build the per-block recall formula from the human evaluation column."""
    if gt_count <= 0:
        return ""
    return f'=COUNTIF(K{start_row}:K{end_row},"CS")/{gt_count}'


def _count_numbered_steps(steps_text: str) -> int:
    """Count numbered S2R lines in the ground-truth CSV format."""
    return sum(
        1
        for line in steps_text.splitlines()
        if line.strip() and line.lstrip().split(".", 1)[0].isdigit()
    )


def _append_separator_row(sheet: Any, column_count: int, row_number: int) -> None:
    """Insert a blank gray spacer row between S2R result blocks."""
    sheet.append([""] * column_count)
    for cell in sheet[row_number]:
        cell.fill = SEPARATOR_FILL


def _append_error(existing_error: str | None, new_error: str) -> str:
    """Accumulate multiple evaluator errors into one readable field."""
    if not existing_error:
        return new_error
    return f"{existing_error} | {new_error}"


if __name__ == "__main__":
    main()
